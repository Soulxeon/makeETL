import os
import settings
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def front(etl, dataset, excluded):

    logo = Image(os.path.join(os.path.dirname(__file__), "img/imago_logo_tiny.png"))
    
    wb = Workbook()
    ws = wb.active

    ws.title = "Front"

    #region front page
    for r in dataframe_to_rows(dataset, index=True, header=True):
        ws.append(r)

    #region Frontstyle
    ws['A{}'.format(ws.max_row)].font = Font(bold=True)

    headspace = ws.max_row
    uploaded = headspace + 3
    etlstart = headspace + 5 

    total_images = ws.cell(row = headspace, column = 3).value

    ws.merge_cells('F4:L5')
    ws.merge_cells('O2:Q8')
    ws.add_image(logo,'O2')
    
    if settings.pre_upload:
        prefixRep = 'PreUpload Report'
        prefixUp = 'to upload'
        prefixExc = 'excluded'   
    else:
        prefixRep = 'PostUpload Report'
        prefixUp = 'uploaded'
        prefixExc = 'excluded'   

    ws.cell(row = 4, column = 6, value = settings.name_project + prefixRep ).font = Font(size = "20")
    ws.cell(row = 4, column = 6).alignment = Alignment(horizontal='center')

    ws.cell(row = uploaded, column = 1, value = prefixUp)
    ws.cell(row = uploaded, column = 2, value = total_images - excluded).font = Font(bold=True, color = '008000')
    ws.cell(row = uploaded, column = 4, value = prefixExc )
    ws.cell(row = uploaded, column = 5, value = excluded).font = Font(bold=True, color = 'FF0000')
    #endregion

    ws.cell(row=etlstart, column = 1) 

    for r in dataframe_to_rows(etl, index=False, header=True):
        ws.append(r)
        
    for y in range(1, 16):
        ws.cell(row=etlstart+1, column=y).fill = PatternFill(start_color='248DE3', end_color='248DE3', fill_type='solid')

    wb.save(settings.report)
